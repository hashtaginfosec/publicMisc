import unicodedata
import openpyxl
import sys

# Open workbooks
workbookv13 = openpyxl.load_workbook('workbookv13.xlsx')
workbookv12 = openpyxl.load_workbook('workbookv12.xlsx')

# Sleect sheets
sheetV13 = workbookv13.get_sheet_by_name("Matrix")
sheetV12 = workbookv12.get_sheet_by_name("Matrix")

fixedRows = []
baselineUniqueIDIssues = []

try:
	for rowNum in range(2, 599):	#for each row
		
		baselineUniqueIDv13 = sheetV13.cell(row=rowNum, column=15).value.decode('utf-8', 'ignore')
		if sheetV13.cell(row=rowNum, column=18).value is not None:
			evidenceReqeustV13 = str(sheetV13.cell(row=rowNum, column=18).value.replace(u'\u2022', u'\-').decode('utf-8', 'ignore'))
		# print(rowNum)
		
		for v12RowNum in range(2, 599):
			if v12RowNum not in fixedRows:
				baselineUniqueIDv12 = str(sheetV12.cell(row=rowNum, column=15).value)
				if sheetV12.cell(row=rowNum, column=18).value is not None:
					evidenceReqeustv12 = str(sheetV12.cell(row=rowNum, column=18).value.replace(u'\u2022', u'\-').decode('utf-8', 'ignore'))
				
				if baselineUniqueIDv13 == baselineUniqueIDv12:
					if evidenceReqeustV13 != evidenceReqeustv12:
						baselineUniqueIDIssues.append(str(baselineUniqueIDv12))
						#make the update
						sheetV13.cell(row=rowNum, column=18).value = evidenceReqeustv12
						fixedRows.append(v12RowNum)
						break
				
except IndexError as e:
    print(e)

#comment
baselineUniqueIDIssues = set(baselineUniqueIDIssues)  #delete duplicates
for item in baselineUniqueIDIssues:
	print(item)
workbookv13.save('workbookv13.xlsx')
print(len(baselineUniqueIDIssues))
